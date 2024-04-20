from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, PatternFill, Border
from openpyxl.drawing.image import Image
import src.analitica as analitica
import pandas as pd
from io import BytesIO

def descargar_excel(wb):
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes


def crear_reporte_excel(year):
    # Crear un nuevo libro y seleccionar la hoja activa
    wb = Workbook()
    ws = wb.active

    # Añadir texto a una celda
    ws['E1'] = f"Reporte Personalizado {year}"

    # Establecer fuente, tamaño y color
    ws['E1'].font = Font(name='American Typewriter', size = 20, bold=True, italic=True,color='005DFF')


    # Ajustar el tamaño de la celda
    ws.row_dimensions[1].height = 30 # Mi letra es de 20
    ws.column_dimensions['E'].width = 60

    # Alinear el texto
    ws['E1'].alignment = Alignment(horizontal='center',vertical='center')

    # Aplicar un color de fondo a una celda
    ws['E1'].fill = PatternFill(start_color='050505',end_color='050505',fill_type = 'solid')

    img = Image(f'Images/donut_platform_year_{year}.png')
    ws.add_image(img,'A4')

    img = Image(f'Images/histogram_age_year_{year}.png')
    ws.add_image(img,'F4')

    plataformas = ['Netflix','Hulu','Prime Video', 'Disney+']
    valores = analitica.plataforma(year)

    edades = analitica.edades_permitidas(year)

    for i in range(37,41):
        ws[f'C{str(i)}'] = plataformas[i-37]
        ws[f'C{str(i)}'].font = Font(name='American Typewriter', size = 12, bold=True, italic=True,color='005DFF')
        ws[f'D{str(i)}'] = valores[plataformas[i-37]]

    for i in range(37,41):
        ws[f'L{str(i)}'] = edades.iloc[i-37]['Age']
        ws[f'L{str(i)}'].font = Font(name='American Typewriter', size = 12, bold=True, italic=True,color='005DFF')
        ws[f'M{str(i)}'] = edades.iloc[i-37]['Amount']

    return wb